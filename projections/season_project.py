#!/usr/bin/env python3
"""
season_project.py — Season-long WNBA fantasy projections
=========================================================
Uses 2020–2024 historical data to project full 2025 season stat lines.
Designed for pre-draft fantasy use: per-game averages + season totals.

Model design:
  1. Aggregate per-player per-season box score rates (per-36 min)
  2. Recency-weighted average across seasons (2024 = 50%, 2020 = 2%)
  3. Apply experience curve (+10% for young players, gradual decline for vets)
  4. Project games played separately via historical GP% × 2025 schedule (45 games)
  5. Convert per-36 → per-game → season totals
  6. Output formatted Excel spreadsheet

Backtest note:
  Training data: 2020–2024  |  Target: 2025 actuals (already in DB)
  Run compare_season.py after to measure accuracy vs 2025 actuals.
"""

import os
import sys
import psycopg2
import pandas as pd
import numpy as np
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

load_dotenv()

# ── Season weights ─────────────────────────────────────────────────────────────
# 2020 is heavily downweighted: bubble season, no travel, unusual lineups/load
SEASON_WEIGHTS = {
    2020: 0.02,
    2021: 0.08,
    2022: 0.15,
    2023: 0.25,
    2024: 0.50,
}

# Actual max team games per season (home + away, regular season only)
MAX_TEAM_GAMES = {
    2020: 23,
    2021: 33,
    2022: 37,
    2023: 42,
    2024: 41,
    2025: 45,
}

PROJECT_GAMES    = 45   # 2025 regular season length
MIN_GAMES        = 5    # Minimum games in a season to include that season
MIN_MINUTES      = 3.0  # Filter out DNPs / garbage time rows

# ── Games-played model constants (derived from 2021–2024 empirical data) ───────
LEAGUE_AVG_GP_PCT = 0.727   # Historical league-wide average GP%
GP_SHRINKAGE      = 2.0     # Equivalent seasons of prior data pulled toward mean
                            # (1 season → 33% individual / 67% mean; 5 seasons → 71%/29%)
INJURY_THRESHOLD  = 0.80    # GP% below this = "injured season"
INJURY_PENALTY    = 0.148   # Empirical: prior injury → 14.8pp lower next-year GP%
TREND_MIN_SEASONS = 3       # Need at least this many seasons to apply trend adjustment

# Stats to model via per-36 rates
RATE_STATS = ['pts', 'reb', 'ast', 'stl', 'blk', 'tov', 'fg3m', 'fgm', 'fga', 'ftm', 'fta']

# Fantasy scoring weights  (stat: points per unit)
SCORING = {
    'Your League': dict(pts=1.0, reb=1.2,  ast=1.5, stl=3.0, blk=3.0, fg3m=0.5, tov=-1.0),
    'ESPN':        dict(pts=1.0, reb=1.0,  ast=2.0, stl=4.0, blk=4.0, fg3m=1.0, tov=-2.0),
    'DraftKings':  dict(pts=1.0, reb=1.25, ast=1.5, stl=2.0, blk=2.0, fg3m=0.5, tov=-0.5),
    'Yahoo':       dict(pts=1.0, reb=1.0,  ast=1.5, stl=3.0, blk=3.0, fg3m=1.0, tov=-1.5),
}

# ── Excel styling ──────────────────────────────────────────────────────────────
HEADER_BG   = '1B4F72'   # dark teal
HEADER_FONT = 'FFFFFF'   # white
ALT_ROW_BG  = 'EAF4FB'   # light blue
SECTION_BG  = 'D4E6F1'   # slightly darker blue for section headers


# ── Database helpers ───────────────────────────────────────────────────────────

def get_conn():
    return psycopg2.connect(os.environ['PG_DSN'])


def fetch_player_seasons(conn):
    """Per-player per-season aggregates for 2020–2024 (training data)."""
    sql = """
        WITH filtered AS (
            SELECT
                gps.player_id,
                g.season,
                g.game_id,
                gps.min_played::float                    AS min_played,
                COALESCE(gps.pts,  0)::float             AS pts,
                COALESCE(gps.reb,  0)::float             AS reb,
                COALESCE(gps.ast,  0)::float             AS ast,
                COALESCE(gps.stl,  0)::float             AS stl,
                COALESCE(gps.blk,  0)::float             AS blk,
                COALESCE(gps.tov,  0)::float             AS tov,
                COALESCE(gps.fg3m, 0)::float             AS fg3m,
                COALESCE(gps.fgm,  0)::float             AS fgm,
                COALESCE(gps.fga,  0)::float             AS fga,
                COALESCE(gps.ftm,  0)::float             AS ftm,
                COALESCE(gps.fta,  0)::float             AS fta
            FROM game_player_stats gps
            JOIN games g ON gps.game_id = g.game_id
            WHERE g.league = 'wnba'
              AND (g.postseason = false OR g.postseason IS NULL)
              AND g.season BETWEEN 2020 AND 2024
              AND gps.min_played >= %s
        )
        SELECT
            player_id,
            season,
            COUNT(DISTINCT game_id)                                    AS games_played,
            SUM(min_played) / COUNT(DISTINCT game_id)                  AS mpg,
            SUM(pts)  / NULLIF(SUM(min_played), 0) * 36               AS pts_per36,
            SUM(reb)  / NULLIF(SUM(min_played), 0) * 36               AS reb_per36,
            SUM(ast)  / NULLIF(SUM(min_played), 0) * 36               AS ast_per36,
            SUM(stl)  / NULLIF(SUM(min_played), 0) * 36               AS stl_per36,
            SUM(blk)  / NULLIF(SUM(min_played), 0) * 36               AS blk_per36,
            SUM(tov)  / NULLIF(SUM(min_played), 0) * 36               AS tov_per36,
            SUM(fg3m) / NULLIF(SUM(min_played), 0) * 36               AS fg3m_per36,
            SUM(fgm)  / NULLIF(SUM(min_played), 0) * 36               AS fgm_per36,
            SUM(fga)  / NULLIF(SUM(min_played), 0) * 36               AS fga_per36,
            SUM(ftm)  / NULLIF(SUM(min_played), 0) * 36               AS ftm_per36,
            SUM(fta)  / NULLIF(SUM(min_played), 0) * 36               AS fta_per36
        FROM filtered
        GROUP BY player_id, season
        HAVING COUNT(DISTINCT game_id) >= %s
        ORDER BY player_id, season
    """
    cur = conn.cursor()
    cur.execute(sql, (MIN_MINUTES, MIN_GAMES))
    cols = [d[0] for d in cur.description]
    return pd.DataFrame(cur.fetchall(), columns=cols)


def fetch_active_players(conn):
    """Players who saw meaningful minutes in 2024 — our 2025 projection candidates."""
    sql = """
        SELECT DISTINCT gps.player_id
        FROM game_player_stats gps
        JOIN games g ON gps.game_id = g.game_id
        WHERE g.league = 'wnba'
          AND g.season = 2024
          AND gps.min_played >= %s
    """
    cur = conn.cursor()
    cur.execute(sql, (MIN_MINUTES,))
    return [r[0] for r in cur.fetchall()]


def fetch_names(conn):
    sql = """
        SELECT player_id,
               TRIM(COALESCE(first_name,'') || ' ' || COALESCE(last_name,'')) AS name
        FROM players
    """
    cur = conn.cursor()
    cur.execute(sql)
    return pd.DataFrame(cur.fetchall(), columns=['player_id', 'name'])


def fetch_archetypes(conn):
    # Use most recent season's archetype per player
    sql = """
        SELECT DISTINCT ON (player_id) player_id, archetype
        FROM player_archetypes
        WHERE league = 'wnba'
        ORDER BY player_id, season DESC
    """
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    return pd.DataFrame(rows, columns=['player_id', 'archetype'])


# ── Projection model ───────────────────────────────────────────────────────────

def weighted_projection(player_hist: pd.DataFrame):
    """
    Compute season-weight-normalized per-36 rates + MPG + GP% for one player.
    Returns None if no usable seasons exist.
    """
    acc = {f'{s}_per36': 0.0 for s in RATE_STATS}
    acc_mpg = 0.0
    acc_gp_pct = 0.0
    total_w = 0.0

    for _, row in player_hist.iterrows():
        season = int(row['season'])
        w = SEASON_WEIGHTS.get(season, 0.0)
        if w == 0:
            continue

        gp_pct = min(float(row['games_played']) / MAX_TEAM_GAMES[season], 1.0)

        for stat in RATE_STATS:
            v = row.get(f'{stat}_per36')
            if pd.notna(v):
                acc[f'{stat}_per36'] += w * float(v)

        if pd.notna(row['mpg']):
            acc_mpg += w * float(row['mpg'])

        acc_gp_pct += w * gp_pct
        total_w += w

    if total_w == 0:
        return None

    proj = {k: v / total_w for k, v in acc.items()}
    proj['mpg']    = acc_mpg / total_w
    proj['gp_pct'] = acc_gp_pct / total_w
    return proj


def project_games_played(player_hist: pd.DataFrame, weighted_gp_pct: float) -> tuple:
    """
    Three-component games-played model:

    1. Bayesian shrinkage — pull the weighted GP% toward the league mean
       proportional to how many seasons of data we have.
       Formula: (n * hist_gp + k * league_mean) / (n + k)
       With k=2: 1 season → 33% individual + 67% mean; 5 seasons → 71% + 29% mean

    2. Recent injury penalty — if the most recent season's GP% < 80%, apply
       a scaled penalty based on how severe the miss was (empirical: 14.8pp avg).

    3. Trend adjustment — if GP% has been declining across the last 3 seasons,
       carry forward half the observed slope as a downward nudge.

    Returns: (projected_games: int, final_gp_pct: float)
    """
    n = len(player_hist)

    # ── 1. Bayesian shrinkage ──────────────────────────────────────────────────
    shrunk = (n * weighted_gp_pct + GP_SHRINKAGE * LEAGUE_AVG_GP_PCT) / (n + GP_SHRINKAGE)

    # ── 2. Recent injury penalty ───────────────────────────────────────────────
    most_recent = player_hist.sort_values('season').iloc[-1]
    recent_season = int(most_recent['season'])
    recent_max = MAX_TEAM_GAMES.get(recent_season, 41)
    recent_gp_pct = min(float(most_recent['games_played']) / recent_max, 1.0)

    injury_adj = 0.0
    if recent_gp_pct < INJURY_THRESHOLD:
        # Scale: a player who played 60% of games gets a bigger penalty than one at 79%
        severity = (INJURY_THRESHOLD - recent_gp_pct) / INJURY_THRESHOLD
        injury_adj = -INJURY_PENALTY * severity

    # ── 3. Trend adjustment ────────────────────────────────────────────────────
    trend_adj = 0.0
    if n >= TREND_MIN_SEASONS:
        recent3 = player_hist.sort_values('season').tail(3)
        gp_pcts = [
            min(float(r['games_played']) / MAX_TEAM_GAMES.get(int(r['season']), 41), 1.0)
            for _, r in recent3.iterrows()
        ]
        # Simple slope: change per season over the last 3 seasons
        slope = (gp_pcts[-1] - gp_pcts[0]) / 2.0
        if slope < -0.05:   # Only act on a meaningful decline (>5pp/season)
            trend_adj = slope * 0.5   # Apply half the slope — don't overfit to noise

    # ── Combine & clip ─────────────────────────────────────────────────────────
    final_gp_pct = float(np.clip(shrunk + injury_adj + trend_adj, 0.15, 1.0))
    proj_games = int(np.clip(round(final_gp_pct * PROJECT_GAMES), 1, PROJECT_GAMES))

    return proj_games, final_gp_pct


def experience_curve(proj: dict, seasons: int) -> dict:
    """
    Multiply per-36 rates by an experience-based adjustment.

    Curve (seasons in league as proxy for age):
      1–2  → +10%  (still developing, likely improving)
      3–5  → ±0%   (peak window)
      6–8  → -3% per season above 5
      9+   → -5% per season above 8  (floor: 65%)
    """
    if seasons <= 2:
        mult = 1.10
    elif seasons <= 5:
        mult = 1.00
    elif seasons <= 8:
        mult = 1.00 - 0.03 * (seasons - 5)
    else:
        mult = 0.91 - 0.05 * (seasons - 8)

    mult = max(mult, 0.65)

    for stat in RATE_STATS:
        key = f'{stat}_per36'
        if key in proj:
            proj[key] *= mult
    return proj


def to_per_game(proj: dict):
    """Scale per-36 rates down to projected MPG."""
    mpg = proj.get('mpg', 20.0)
    out = {'proj_min': round(mpg, 1)}
    for stat in RATE_STATS:
        val = proj.get(f'{stat}_per36', 0.0)
        out[f'proj_{stat}'] = round(val * mpg / 36.0, 2)
    return out


def fantasy_pg(row: pd.Series, weights: dict) -> float:
    return round(sum(row.get(f'proj_{s}', 0) * w for s, w in weights.items()), 2)


# ── Build projection DataFrame ─────────────────────────────────────────────────

def build_projections() -> pd.DataFrame:
    conn = get_conn()
    print("Pulling historical data (2020–2024)...")
    hist       = fetch_player_seasons(conn)
    active     = fetch_active_players(conn)
    names      = fetch_names(conn)
    archetypes = fetch_archetypes(conn)
    conn.close()

    print(f"  {hist['player_id'].nunique():,} players with historical seasons")
    print(f"  {len(active):,} players active in 2024 (projection candidates)")

    rows, skipped = [], 0
    for pid in active:
        phist = hist[hist['player_id'] == pid].sort_values('season')
        if phist.empty:
            skipped += 1
            continue

        seasons = len(phist)
        proj = weighted_projection(phist)
        if proj is None:
            skipped += 1
            continue

        proj = experience_curve(proj, seasons)
        pg   = to_per_game(proj)

        proj_games, final_gp_pct = project_games_played(phist, proj['gp_pct'])

        rows.append({'player_id': pid, 'seasons_in_data': seasons,
                     'proj_games': proj_games, 'proj_gp_pct': round(final_gp_pct, 3),
                     **pg})

    df = pd.DataFrame(rows)

    # Attach names / archetypes
    df = df.merge(names,      on='player_id', how='left')
    df = df.merge(archetypes, on='player_id', how='left')
    df['name']      = df.apply(
        lambda r: r['name'] if pd.notna(r.get('name')) and str(r.get('name')).strip()
                  else f"Player #{r['player_id']}", axis=1)
    df['archetype'] = df['archetype'].fillna('Unknown')

    # Per-game fantasy scores
    for label, weights in SCORING.items():
        col = label.lower().replace(' ', '_') + '_pg'
        df[col] = df.apply(lambda r: fantasy_pg(r, weights), axis=1)

    # Season totals
    for stat in RATE_STATS:
        df[f'season_{stat}'] = (df[f'proj_{stat}'] * df['proj_games']).round(1)

    for label in SCORING:
        pg_col  = label.lower().replace(' ', '_') + '_pg'
        sea_col = label.lower().replace(' ', '_') + '_season'
        df[sea_col] = (df[pg_col] * df['proj_games']).round(1)

    df = df.sort_values('your_league_season', ascending=False).reset_index(drop=True)

    print(f"\n  Projected {len(df)} players  |  {skipped} skipped (no usable history)")
    return df


# ── Excel export ───────────────────────────────────────────────────────────────

def _header_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name='Arial', bold=True, color=HEADER_FONT, size=10)
    c.fill      = PatternFill('solid', fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _data_cell(ws, row, col, value, fmt=None, alt=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Arial', size=9)
    if alt:
        c.fill = PatternFill('solid', fgColor=ALT_ROW_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    if fmt:
        c.number_format = fmt
    return c


def export_excel(df: pd.DataFrame, path: str):
    wb = Workbook()

    # ── Sheet 1: Season Projections ────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Season Projections'
    ws.freeze_panes = 'D2'   # freeze name + archetype + rank

    # Column layout
    # A: Rank  B: Player  C: Archetype  D: Proj Games
    # E: Proj Min  F: Pts  G: Reb  H: Ast  I: Stl  J: Blk  K: Tov  L: 3PM
    # M: YL/game  N: ESPN/game  O: DK/game  P: Yahoo/game
    # Q: YL Season  R: ESPN Season  S: DK Season  T: Yahoo Season

    headers = [
        ('Rank',          5),
        ('Player',        22),
        ('Archetype',     14),
        ('Proj\nGames',   8),
        ('Proj\nMin',     8),
        ('Proj\nPts',     8),
        ('Proj\nReb',     8),
        ('Proj\nAst',     8),
        ('Proj\nStl',     8),
        ('Proj\nBlk',     8),
        ('Proj\nTov',     8),
        ('Proj\n3PM',     8),
        ('Your League\n/Game',  11),
        ('ESPN\n/Game',         9),
        ('DraftKings\n/Game',   10),
        ('Yahoo\n/Game',        9),
        ('Your League\nSeason', 13),
        ('ESPN\nSeason',        11),
        ('DraftKings\nSeason',  12),
        ('Yahoo\nSeason',       11),
    ]

    ws.row_dimensions[1].height = 36
    for col_idx, (label, width) in enumerate(headers, start=1):
        _header_cell(ws, 1, col_idx, label, width)

    # Section divider styling for per-game vs season columns
    section_fill = PatternFill('solid', fgColor=SECTION_BG)

    num_fmt   = '0.0'
    score_fmt = '0.00'
    int_fmt   = '0'

    for i, row in df.iterrows():
        r = i + 2
        alt = (i % 2 == 1)

        vals = [
            (i + 1,                    int_fmt),
            (row['name'],              None),
            (row['archetype'],         None),
            (int(row['proj_games']),   int_fmt),
            (row['proj_min'],          num_fmt),
            (row['proj_pts'],          num_fmt),
            (row['proj_reb'],          num_fmt),
            (row['proj_ast'],          num_fmt),
            (row['proj_stl'],          num_fmt),
            (row['proj_blk'],          num_fmt),
            (row['proj_tov'],          num_fmt),
            (row['proj_fg3m'],         num_fmt),
            (row['your_league_pg'],    score_fmt),
            (row['espn_pg'],           score_fmt),
            (row['draftkings_pg'],     score_fmt),
            (row['yahoo_pg'],          score_fmt),
            (row['your_league_season'],num_fmt),
            (row['espn_season'],       num_fmt),
            (row['draftkings_season'], num_fmt),
            (row['yahoo_season'],      num_fmt),
        ]

        for col_idx, (val, fmt) in enumerate(vals, start=1):
            c = _data_cell(ws, r, col_idx, val, fmt, alt)
            if col_idx == 2:   # Player name left-aligned
                c.alignment = Alignment(horizontal='left', vertical='center')

    n_rows = len(df)
    last_data_row = n_rows + 1

    # Color scale on scoring columns (white → green, low → high)
    score_cols = [13, 14, 15, 16, 17, 18, 19, 20]
    for col_idx in score_cols:
        col_letter = get_column_letter(col_idx)
        cell_range = f'{col_letter}2:{col_letter}{last_data_row}'
        ws.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type='min', start_color='FFFFFF',
                end_type='max',   end_color='00B050',
            )
        )

    # ── Sheet 2: Per-Season Detail ─────────────────────────────────────────────
    ws2 = wb.create_sheet('Per-Season Stats')
    ws2.freeze_panes = 'D2'

    hdrs2 = [
        ('Rank', 5), ('Player', 22), ('Archetype', 14),
        ('Proj\nGames', 8),
        ('Season\nPts', 9), ('Season\nReb', 9), ('Season\nAst', 9),
        ('Season\nStl', 9), ('Season\nBlk', 9), ('Season\nTov', 9),
        ('Season\n3PM', 9), ('Season\nFGM', 9), ('Season\nFGA', 9),
        ('Season\nFTM', 9), ('Season\nFTA', 9),
    ]
    ws2.row_dimensions[1].height = 36
    for col_idx, (label, width) in enumerate(hdrs2, start=1):
        _header_cell(ws2, 1, col_idx, label, width)

    for i, row in df.iterrows():
        r = i + 2
        alt = (i % 2 == 1)
        vals2 = [
            (i + 1,                       int_fmt),
            (row['name'],                 None),
            (row['archetype'],            None),
            (int(row['proj_games']),      int_fmt),
            (row['season_pts'],           num_fmt),
            (row['season_reb'],           num_fmt),
            (row['season_ast'],           num_fmt),
            (row['season_stl'],           num_fmt),
            (row['season_blk'],           num_fmt),
            (row['season_tov'],           num_fmt),
            (row['season_fg3m'],          num_fmt),
            (row['season_fgm'],           num_fmt),
            (row['season_fga'],           num_fmt),
            (row['season_ftm'],           num_fmt),
            (row['season_fta'],           num_fmt),
        ]
        for col_idx, (val, fmt) in enumerate(vals2, start=1):
            c = _data_cell(ws2, r, col_idx, val, fmt, alt)
            if col_idx == 2:
                c.alignment = Alignment(horizontal='left', vertical='center')

    # Color scale on season stat columns
    for col_idx in range(5, 16):
        col_letter = get_column_letter(col_idx)
        cell_range = f'{col_letter}2:{col_letter}{last_data_row}'
        ws2.conditional_formatting.add(
            cell_range,
            ColorScaleRule(
                start_type='min', start_color='FFFFFF',
                end_type='max',   end_color='00B050',
            )
        )

    wb.save(path)
    print(f"\nSaved → {path}")


# ── Accuracy comparison (projected vs actual 2025) ────────────────────────────

def fetch_actual_2025(conn):
    """Pull actual 2025 regular-season totals per player."""
    sql = """
        SELECT
            gps.player_id,
            COUNT(DISTINCT g.game_id)              AS actual_games,
            SUM(COALESCE(gps.pts,  0))::float      AS actual_pts,
            SUM(COALESCE(gps.reb,  0))::float      AS actual_reb,
            SUM(COALESCE(gps.ast,  0))::float      AS actual_ast,
            SUM(COALESCE(gps.stl,  0))::float      AS actual_stl,
            SUM(COALESCE(gps.blk,  0))::float      AS actual_blk,
            SUM(COALESCE(gps.tov,  0))::float      AS actual_tov,
            SUM(COALESCE(gps.fg3m, 0))::float      AS actual_fg3m
        FROM game_player_stats gps
        JOIN games g ON gps.game_id = g.game_id
        WHERE g.league = 'wnba'
          AND g.season = 2025
          AND (g.postseason = false OR g.postseason IS NULL)
          AND gps.min_played >= 3
        GROUP BY gps.player_id
        HAVING COUNT(DISTINCT g.game_id) >= 5
    """
    cur = conn.cursor()
    cur.execute(sql)
    cols = [d[0] for d in cur.description]
    return pd.DataFrame(cur.fetchall(), columns=cols)


def build_accuracy_df(proj_df: pd.DataFrame, actuals: pd.DataFrame) -> pd.DataFrame:
    """
    Join projections with 2025 actuals. Compute actual fantasy scores under each
    scoring system, then measure projected vs actual for each format.
    """
    df = proj_df.merge(actuals, on='player_id', how='inner')

    # Actual fantasy season scores under each format
    for label, weights in SCORING.items():
        col = label.lower().replace(' ', '_') + '_actual'
        df[col] = (
            df['actual_pts']  * weights.get('pts',  0) +
            df['actual_reb']  * weights.get('reb',  0) +
            df['actual_ast']  * weights.get('ast',  0) +
            df['actual_stl']  * weights.get('stl',  0) +
            df['actual_blk']  * weights.get('blk',  0) +
            df['actual_tov']  * weights.get('tov',  0) +
            df['actual_fg3m'] * weights.get('fg3m', 0)
        ).round(1)

    # Error = projected season score − actual season score
    for label in SCORING:
        slug = label.lower().replace(' ', '_')
        df[f'{slug}_error']   = (df[f'{slug}_season'] - df[f'{slug}_actual']).round(1)
        df[f'{slug}_abs_err'] = df[f'{slug}_error'].abs()

    # Actual rank vs projected rank (using Your League)
    df['actual_rank'] = df['your_league_actual'].rank(ascending=False, method='min').astype(int)
    df['proj_rank']   = df['your_league_season'].rank(ascending=False, method='min').astype(int)
    df['rank_diff']   = df['proj_rank'] - df['actual_rank']   # negative = we ranked them too high

    return df.sort_values('your_league_actual', ascending=False).reset_index(drop=True)


def _accuracy_header(ws, row, col, value, width=None, bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name='Arial', bold=True, color=HEADER_FONT, size=10)
    c.fill      = PatternFill('solid', fgColor=bg or HEADER_BG)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def add_accuracy_sheet(wb, acc_df: pd.DataFrame):
    """
    Sheet 3: Projected vs Actual — one row per player, columns grouped by format.

    Layout:
      A  Actual Rank    B  Player           C  Archetype      D  Proj Rank   E  Rank Diff
      F  Actual Games   G  Projected Games

      Then for each scoring format (Your League, ESPN, DraftKings, Yahoo):
        Projected Season | Actual Season | Difference
    """
    ws = wb.create_sheet('Projected vs Actual')
    ws.freeze_panes = 'C2'
    ws.row_dimensions[1].height = 40

    # Section header colors per format
    format_colors = {
        'Your League': '1A5276',   # dark blue
        'ESPN':        '145A32',   # dark green
        'DraftKings':  '6E2F2F',   # dark red
        'Yahoo':       '4A235A',   # dark purple
    }
    format_light = {
        'Your League': 'D6EAF8',
        'ESPN':        'D5F5E3',
        'DraftKings':  'FADBD8',
        'Yahoo':       'E8DAEF',
    }

    # Build column definitions
    fixed_cols = [
        ('Actual\nRank',      6),
        ('Player',           22),
        ('Archetype',        13),
        ('Projected\nRank',   8),
        ('Rank\nDiff',        7),
        ('Actual\nGames',     7),
        ('Projected\nGames',  8),
    ]

    # Write fixed headers
    for col_idx, (label, width) in enumerate(fixed_cols, start=1):
        _accuracy_header(ws, 1, col_idx, label, width)

    # Write format-grouped headers (3 cols each: Proj Season | Actual Season | Diff)
    col = len(fixed_cols) + 1
    format_col_starts = {}
    for label in SCORING:
        format_col_starts[label] = col
        color = format_colors[label]

        _accuracy_header(ws, 1, col,     f'{label}\nProjected', 13, color)
        _accuracy_header(ws, 1, col + 1, f'{label}\nActual',    13, color)
        _accuracy_header(ws, 1, col + 2, f'{label}\nDiff',      10, color)
        col += 3

    # Data rows
    num_fmt   = '#,##0.0'
    int_fmt   = '0'
    diff_fmt  = '+#,##0.0;-#,##0.0;0.0'
    rank_diff = '[Blue]+0;[Red]-0;0'

    for i, row in acc_df.iterrows():
        r = i + 2
        alt = (i % 2 == 1)

        def dc(col_idx, val, fmt=None, color_override=None):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = Font(name='Arial', size=9)
            if color_override:
                c.fill = PatternFill('solid', fgColor=color_override)
            elif alt:
                c.fill = PatternFill('solid', fgColor=ALT_ROW_BG)
            c.alignment = Alignment(horizontal='center', vertical='center')
            if fmt:
                c.number_format = fmt
            return c

        dc(1, int(row['actual_rank']),  int_fmt)
        c = dc(2, row['name'])
        c.alignment = Alignment(horizontal='left', vertical='center')
        dc(3, row['archetype'])
        dc(4, int(row['proj_rank']),    int_fmt)
        dc(5, int(row['rank_diff']),    rank_diff)
        dc(6, int(row['actual_games']), int_fmt)
        dc(7, int(row['proj_games']),   int_fmt)

        col = len(fixed_cols) + 1
        for label in SCORING:
            slug       = label.lower().replace(' ', '_')
            light_col  = format_light[label]
            proj_val   = row[f'{slug}_season']
            actual_val = row[f'{slug}_actual']
            diff_val   = row[f'{slug}_error']

            dc(col,     proj_val,   num_fmt,  light_col if not alt else None)
            dc(col + 1, actual_val, num_fmt,  light_col if not alt else None)
            dc(col + 2, diff_val,   diff_fmt, light_col if not alt else None)
            col += 3

    last_row = len(acc_df) + 1

    # Color scale on Diff columns (red=over, white=0, green=under)
    col = len(fixed_cols) + 3   # first Diff column
    for label in SCORING:
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{last_row}',
            ColorScaleRule(
                start_type='min',  start_color='00B050',   # green = underestimated
                mid_type='num',    mid_value=0, mid_color='FFFFFF',
                end_type='max',    end_color='FF0000',      # red = overestimated
            )
        )
        col += 3

    # ── Summary stats block (below data) ──────────────────────────────────────
    summary_row = last_row + 3
    ws.cell(row=summary_row, column=1, value='Accuracy Summary (vs 2025 Actuals)').font = \
        Font(name='Arial', bold=True, size=11)

    ws.cell(row=summary_row + 1, column=1, value='Format').font = Font(name='Arial', bold=True)
    for i, metric in enumerate(['MAE (Fantasy Pts)', 'Avg Error (Bias)', 'RMSE', 'Within 50 pts %', 'Within 100 pts %'], start=2):
        ws.cell(row=summary_row + 1, column=i, value=metric).font = Font(name='Arial', bold=True)

    for r_off, label in enumerate(SCORING, start=2):
        slug = label.lower().replace(' ', '_')
        errors = acc_df[f'{slug}_error']
        mae    = errors.abs().mean()
        bias   = errors.mean()
        rmse   = np.sqrt((errors ** 2).mean())
        w50    = (errors.abs() <= 50).mean() * 100
        w100   = (errors.abs() <= 100).mean() * 100

        ws.cell(row=summary_row + r_off, column=1, value=label).font = Font(name='Arial', bold=True)
        for c_off, val in enumerate([mae, bias, rmse, w50, w100], start=2):
            cell = ws.cell(row=summary_row + r_off, column=c_off, value=round(val, 1))
            cell.font = Font(name='Arial', size=9)
            if c_off in (4, 5):  # pct columns
                cell.number_format = '0.0"%"'

    # Print to console too
    print('\n── Accuracy vs 2025 Actuals ──')
    print(f'  {"Format":<14} {"MAE":>8} {"Bias":>8} {"RMSE":>8} {"≤50pts":>8} {"≤100pts":>8}')
    for label in SCORING:
        slug   = label.lower().replace(' ', '_')
        errors = acc_df[f'{slug}_error']
        mae    = errors.abs().mean()
        bias   = errors.mean()
        rmse   = np.sqrt((errors ** 2).mean())
        w50    = (errors.abs() <= 50).mean() * 100
        w100   = (errors.abs() <= 100).mean() * 100
        print(f'  {label:<14} {mae:>8.1f} {bias:>+8.1f} {rmse:>8.1f} {w50:>7.1f}% {w100:>7.1f}%')

    print('\n── Biggest misses (Your League, by absolute error) ──')
    top_miss = acc_df.nlargest(8, 'your_league_abs_err')[
        ['name', 'proj_rank', 'actual_rank', 'rank_diff',
         'proj_games', 'actual_games',
         'your_league_season', 'your_league_actual', 'your_league_error']
    ]
    print(top_miss.to_string(index=False))


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    conn = get_conn()
    actuals = fetch_actual_2025(conn)
    conn.close()

    df = build_projections()

    # Accuracy comparison
    acc_df = build_accuracy_df(df, actuals)

    # Console preview
    print("\n── Top 20: Season Projections (sorted by Your League Season) ──")
    preview = df[['name', 'archetype', 'proj_games', 'proj_min', 'proj_pts',
                  'proj_reb', 'proj_ast', 'proj_stl', 'proj_blk',
                  'your_league_pg', 'your_league_season', 'espn_season']].head(20)
    print(preview.to_string(index=False))

    out_path = os.path.join(os.path.dirname(__file__), 'season_projections_2025.xlsx')
    export_excel(df, out_path)

    # Load workbook back and add accuracy sheet
    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    add_accuracy_sheet(wb, acc_df)
    wb.save(out_path)
    print(f"Accuracy sheet added → {out_path}")

    return df, acc_df


if __name__ == '__main__':
    main()
